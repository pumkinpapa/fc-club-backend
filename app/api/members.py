"""
회원 관리 API
"""

from typing import List

from fastapi import APIRouter, Depends, HTTPException, Query
from sqlalchemy.orm import Session

from app.core.database import get_db
from app.core.security import get_current_user, get_admin_user
from app.models import Member
from app.schemas import MemberResponse, MemberUpdate
# ★ 신규: 시스템관리자 권한 + cascade 삭제 서비스
from app.services.match_service import delete_member_cascade

router = APIRouter(prefix="/api/members", tags=["회원관리"])


# 시스템관리자 전화번호 (matches.py와 동일하게 유지)
SYS_ADMIN_PHONE = "01000000001"


def get_system_admin_user(current_user: Member = Depends(get_current_user)) -> Member:
    """시스템관리자 권한 확인"""
    if current_user.phone != SYS_ADMIN_PHONE:
        raise HTTPException(status_code=403, detail="시스템관리자만 수행 가능합니다.")
    return current_user


# ──────────────────────────────────
# 회원 조회
# ──────────────────────────────────

@router.get("/", response_model=List[MemberResponse])
async def list_members(
    db: Session = Depends(get_db),
    current_user: Member = Depends(get_current_user),
    status: str = Query("승인", description="회원 상태 필터"),
):
    query = db.query(Member).order_by(Member.join_date)
    if status != "전체":
        query = query.filter(Member.status == status)
    members = query.all()
    return [MemberResponse.model_validate(m) for m in members]


@router.get("/pending", response_model=List[MemberResponse])
async def list_pending_members(
    db: Session = Depends(get_db),
    admin: Member = Depends(get_admin_user),
):
    members = db.query(Member).filter(Member.status == "대기").order_by(Member.join_date).all()
    return [MemberResponse.model_validate(m) for m in members]


@router.post("/{member_id}/approve")
async def approve_member(
    member_id: int,
    db: Session = Depends(get_db),
    admin: Member = Depends(get_admin_user),
):
    member = db.query(Member).filter(Member.id == member_id).first()
    if not member:
        raise HTTPException(status_code=404, detail="회원을 찾을 수 없습니다.")
    member.status = "승인"
    db.commit()
    db.refresh(member)
    return {"message": f"{member.name}님이 승인되었습니다."}


@router.post("/{member_id}/reject")
async def reject_member(
    member_id: int,
    db: Session = Depends(get_db),
    admin: Member = Depends(get_admin_user),
):
    """
    가입 신청 거절 = 회원 기록 완전 삭제

    - 기존에는 status="거절"로 마킹만 했으나, 투표 탭의 '미응답'에 계속 남는 문제가 있음
    - 거절 시 Member 및 관련 MatchRecord를 완전 삭제하여 모든 화면에서 즉시 사라지도록 변경
    """
    try:
        name = delete_member_cascade(db, member_id)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    return {"message": f"{name}님의 가입 신청이 거절되었습니다."}


@router.get("/{member_id}", response_model=MemberResponse)
async def get_member(
    member_id: int,
    db: Session = Depends(get_db),
    current_user: Member = Depends(get_current_user),
):
    member = db.query(Member).filter(Member.id == member_id).first()
    if not member:
        raise HTTPException(status_code=404, detail="회원을 찾을 수 없습니다.")
    return MemberResponse.model_validate(member)


@router.put("/{member_id}", response_model=MemberResponse)
async def update_member(
    member_id: int,
    update: MemberUpdate,
    db: Session = Depends(get_db),
    admin: Member = Depends(get_admin_user),
):
    member = db.query(Member).filter(Member.id == member_id).first()
    if not member:
        raise HTTPException(status_code=404, detail="회원을 찾을 수 없습니다.")
    update_data = update.model_dump(exclude_unset=True)
    for key, value in update_data.items():
        setattr(member, key, value)
    db.commit()
    db.refresh(member)
    return MemberResponse.model_validate(member)


# ══════════════════════════════════════════════
# ★★★ 신규: 선호 포지션 업데이트 (본인만) ★★★
# ══════════════════════════════════════════════

from pydantic import BaseModel
from typing import Optional

class PositionsUpdate(BaseModel):
    positions: str  # "ST,CM" 형식 (쉼표 구분, 최대 2개), 빈 문자열이면 초기화


class PhotoUpdate(BaseModel):
    photo: str  # Base64 data URL (예: "data:image/jpeg;base64,...") 또는 빈 문자열


VALID_POSITIONS = {"GK","CB","LB","RB","DM","CM","AM","LM","RM","ST","LW","RW"}


@router.put("/{member_id}/positions", response_model=MemberResponse)
async def update_positions(
    member_id: int,
    update: PositionsUpdate,
    db: Session = Depends(get_db),
    current_user: Member = Depends(get_current_user),
):
    """
    선호 포지션 업데이트 (본인만 수정 가능)
    - 최대 2개 포지션 (쉼표 구분)
    - 빈 문자열이면 초기화
    - 유효한 포지션 코드만 허용
    """
    member = db.query(Member).filter(Member.id == member_id).first()
    if not member:
        raise HTTPException(status_code=404, detail="회원을 찾을 수 없습니다.")

    # 본인만 수정 가능 (시스템관리자는 예외)
    if current_user.id != member_id and current_user.phone != SYS_ADMIN_PHONE:
        raise HTTPException(status_code=403, detail="본인의 포지션만 수정할 수 있습니다.")

    # 유효성 검사
    positions_str = (update.positions or "").strip()
    if positions_str:
        positions_list = [p.strip().upper() for p in positions_str.split(",") if p.strip()]
        if len(positions_list) > 2:
            raise HTTPException(status_code=400, detail="포지션은 최대 2개까지 선택할 수 있습니다.")
        invalid = [p for p in positions_list if p not in VALID_POSITIONS]
        if invalid:
            raise HTTPException(status_code=400, detail=f"유효하지 않은 포지션: {', '.join(invalid)}")
        positions_str = ",".join(positions_list)

    member.positions = positions_str
    db.commit()
    db.refresh(member)
    return MemberResponse.model_validate(member)


# ══════════════════════════════════════════════
# ★★★ 신규: 프로필 사진 업로드/삭제 (본인 또는 시스템관리자) ★★★
# ══════════════════════════════════════════════

PHOTO_MAX_SIZE_KB = 200  # 200KB 제한 (Base64 기준, 리사이즈 후 대부분 50KB 이하)


@router.put("/{member_id}/photo", response_model=MemberResponse)
async def update_photo(
    member_id: int,
    update: PhotoUpdate,
    db: Session = Depends(get_db),
    current_user: Member = Depends(get_current_user),
):
    """
    프로필 사진 업데이트 (본인 또는 시스템관리자)
    - Base64 data URL 형식
    - 빈 문자열이면 삭제
    - 클라이언트에서 미리 리사이즈 필요 (256x256 권장)
    - 200KB 초과 시 거부
    """
    member = db.query(Member).filter(Member.id == member_id).first()
    if not member:
        raise HTTPException(status_code=404, detail="회원을 찾을 수 없습니다.")

    # 본인만 수정 가능 (시스템관리자는 예외)
    if current_user.id != member_id and current_user.phone != SYS_ADMIN_PHONE:
        raise HTTPException(status_code=403, detail="본인의 사진만 수정할 수 있습니다.")

    photo_data = (update.photo or "").strip()

    # 빈 문자열이면 삭제
    if not photo_data:
        member.photo = ""
        db.commit()
        db.refresh(member)
        return MemberResponse.model_validate(member)

    # Data URL 형식 검증
    if not photo_data.startswith("data:image/"):
        raise HTTPException(
            status_code=400,
            detail="유효하지 않은 이미지 형식입니다. (data:image/... 형식 필요)",
        )

    # 크기 제한
    size_kb = len(photo_data) / 1024
    if size_kb > PHOTO_MAX_SIZE_KB:
        raise HTTPException(
            status_code=400,
            detail=f"이미지 크기가 너무 큽니다. ({size_kb:.0f}KB > {PHOTO_MAX_SIZE_KB}KB)",
        )

    member.photo = photo_data
    db.commit()
    db.refresh(member)
    return MemberResponse.model_validate(member)


# ══════════════════════════════════════════════
# ★★★ 수정: 회원 삭제 - 시스템관리자 전용 + cascade ★★★
# ══════════════════════════════════════════════

@router.delete("/{member_id}")
async def delete_member(
    member_id: int,
    db: Session = Depends(get_db),
    sys_admin: Member = Depends(get_system_admin_user),  # ★ 시스템관리자로 변경
):
    """
    회원 완전 삭제 (시스템관리자 전용)

    - 회원의 모든 MatchRecord(투표/팀/결과) 삭제
    - Member 레코드 삭제
    - 시스템관리자 자신은 삭제 불가
    - 삭제와 동시에 랭킹에서 제거됨 (실시간 집계)
    """
    try:
        name = delete_member_cascade(db, member_id)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    return {"message": f"{name}님이 삭제되었습니다."}


# ══════════════════════════════════════════════
# ★★★ 신규: 회원 정보 엑셀 다운로드 (시스템관리자) ★★★
# ══════════════════════════════════════════════

from fastapi.responses import StreamingResponse
from io import BytesIO
from datetime import datetime


@router.get("/export/excel")
async def export_members_excel(
    db: Session = Depends(get_db),
    sys_admin: Member = Depends(get_system_admin_user),
):
    """
    회원 정보 엑셀 다운로드 (시스템관리자 전용)

    - 승인된 회원 전체
    - 이름, 생년월일, 전화번호, 가입일, 역할, 선호 포지션
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="openpyxl 라이브러리가 설치되지 않았습니다. requirements.txt에 openpyxl 추가 필요.",
        )

    members = db.query(Member).filter(Member.status == "승인").order_by(Member.join_date).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "FC 서울숲 회원"

    # 헤더 스타일
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="34D399", end_color="34D399", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    # 헤더 작성
    headers = ["번호", "이름", "생년월일", "전화번호", "가입일", "역할", "선호 포지션", "상태"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # 데이터 작성
    for row_idx, member in enumerate(members, 2):
        join_str = member.join_date.strftime("%Y-%m-%d") if member.join_date else ""
        birth_str = member.birth or ""
        role = member.role or "회원"
        positions = member.positions or ""

        data = [
            row_idx - 1,  # 번호
            member.name,
            birth_str,
            member.phone,
            join_str,
            role,
            positions,
            member.status,
        ]
        for col_idx, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

    # 컬럼 너비 자동 조정
    column_widths = [8, 14, 14, 16, 14, 10, 14, 10]
    for col_idx, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    # 헤더 행 높이
    ws.row_dimensions[1].height = 24

    # 필터 추가
    ws.auto_filter.ref = ws.dimensions

    # 바이트 스트림으로 저장
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    today = datetime.now().strftime("%Y%m%d")
    filename = f"FC서울숲_회원목록_{today}.xlsx"
    # 한글 파일명 인코딩 (RFC 5987)
    from urllib.parse import quote
    encoded_filename = quote(filename)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}",
        },
    )
